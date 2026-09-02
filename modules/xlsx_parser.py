# -*- coding: utf-8 -*-
"""
XLSX 비교표 추출 모듈 — openpyxl 로 시트별 표를 읽고 병합 셀(merged_cells) 정보를 보존한다
(명세서 6장 STEP 3, 8장).

반환 형식:
{
  "source": "xlsx",
  "slide": "{시트명}",        # 시트명을 slide 필드에 기록
  "table_index": 0,           # 시트 내 표 순번 (여기서는 시트 단위)
  "rows": [ [ {"text","rowspan","colspan"}, ... ], ... ]
}
"""
import io

from openpyxl import load_workbook


class XlsxParseError(Exception):
    """XLSX 읽기/파싱 오류 (사용자에게 그대로 노출)"""


def _cell_value(cell):
    if cell is None:
        return ""
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def parse_xlsx(file_bytes):
    """
    XLSX 파일 바이트를 받아 첫 번째 시트의 표를 추출한다.
    병합 영역은 시작 셀(좌상단)에 rowspan/colspan 을 기록하고,
    병합된 나머지 셀은 빈 문자열로 채운다(구조 보존).
    """
    if not file_bytes:
        raise XlsxParseError("빈 파일입니다. XLSX 파일을 선택해 주세요.")
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    except Exception as exc:
        raise XlsxParseError(f"XLSX 파일을 열지 못했습니다. 형식을 확인해 주세요. ({exc})")
    if not wb.sheetnames:
        raise XlsxParseError("XLSX 파일에 시트가 없습니다.")
    ws = wb[wb.sheetnames[0]]

    nrows, ncols = ws.max_row or 0, ws.max_column or 0
    if nrows == 0 or ncols == 0:
        raise XlsxParseError(f"'{ws.title}' 시트에 데이터가 없습니다.")

    grid = [[{"text": "", "rowspan": 1, "colspan": 1} for _ in range(ncols)] for _ in range(nrows)]
    for r in range(1, nrows + 1):
        for c in range(1, ncols + 1):
            grid[r - 1][c - 1]["text"] = _cell_value(ws.cell(row=r, column=c))
    # 병합 셀 정보 반영
    for rng in ws.merged_cells.ranges:
        top = max(1, rng.min_row) - 1
        left = max(1, rng.min_col) - 1
        h = rng.max_row - rng.min_row + 1
        w = rng.max_col - rng.min_col + 1
        if 0 <= top < nrows and 0 <= left < ncols:
            grid[top][left]["rowspan"] = h
            grid[top][left]["colspan"] = w
            for dr in range(h):
                for dc in range(w):
                    if dr or dc:
                        rr, cc = top + dr, left + dc
                        if rr < nrows and cc < ncols:
                            grid[rr][cc]["text"] = ""  # 병합된 위치는 빈 값 (시작 셀에 값 보존)
    rows = [grid[r][:ncols] for r in range(nrows)]
    wb.close()
    return [{
        "source": "xlsx",
        "slide": ws.title,
        "table_index": 0,
        "rows": rows,
    }]
